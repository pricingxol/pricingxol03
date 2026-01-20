import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ----------------------------------------------------
# KONFIGURASI DASAR
# ----------------------------------------------------
st.set_page_config(page_title="XoL Reinstatement (Simple)", layout="wide", page_icon="📊")
st.title("Pricing Excess of Loss dengan Reinstatement 📊")

MAX_ROWS_DISPLAY = 500  # batas tampilan tabel di UI supaya browser tidak berat


# ----------------------------------------------------
# FUNGSI INPUT ANGKA DENGAN PEMISAH RIBUAN
# ----------------------------------------------------
def rupiah_input(label, key, value=0, step=1_000_000_000, min_value=0):
    """
    Input angka tapi tampilan pakai pemisah ribuan, misal 7.500.000.000.
    Backend tetap menyimpan integer.
    """
    if key + "_raw" not in st.session_state:
        st.session_state[key + "_raw"] = int(value)

    col_minus, col_main, col_plus = st.columns([1, 6, 1])

    # tombol minus
    with col_minus:
        if st.button("−", key=f"{key}_minus"):
            st.session_state[key + "_raw"] = max(min_value, st.session_state[key + "_raw"] - step)

    # kotak teks utama
    with col_main:
        formatted = f"{st.session_state[key + '_raw']:,}".replace(",", ".")
        txt = st.text_input(label, value=formatted, key=f"{key}_text")
        cleaned = txt.replace(".", "").replace(",", "")
        try:
            val = int(cleaned)
            if val >= min_value:
                st.session_state[key + "_raw"] = val
        except ValueError:
            # kalau salah ketik, abaikan dan pakai nilai sebelumnya
            pass

    # tombol plus
    with col_plus:
        if st.button("+", key=f"{key}_plus"):
            st.session_state[key + "_raw"] = st.session_state[key + "_raw"] + step

    return st.session_state[key + "_raw"]


# ----------------------------------------------------
# FUNGSI – SIMULASI FREKUENSI & SEVERITAS (NUMPY ONLY)
# ----------------------------------------------------
def simulasi_monte_carlo(jumlah_iterasi, dist_frek, param_frek, data_severitas, seed=42):
    rng = np.random.default_rng(seed)
    data_tabel = []

    for i in range(jumlah_iterasi):
        # frekuensi
        if dist_frek == "poisson":
            mu = param_frek["mu"]
            freq = int(rng.poisson(mu))
        elif dist_frek == "nbinom":
            n = param_frek["n"]
            p = param_frek["p"]
            freq = int(rng.negative_binomial(n, p))
        else:  # geom
            p = param_frek["p"]
            freq = int(rng.geometric(p))

        if freq < 0:
            freq = 0

        if freq > 0:
            sev = rng.choice(data_severitas, size=freq, replace=True)
        else:
            sev = np.array([])

        for j in range(freq):
            data_tabel.append(
                {
                    "Iterasi": f"{i + 1}.{j + 1}",
                    "Severitas": int(sev[j]),
                    "Flagging Frekuensi": freq if j == 0 else None,
                }
            )

    if not data_tabel:
        raise ValueError("Tidak ada klaim yang dihasilkan dari simulasi (freq = 0 semua).")

    return pd.DataFrame(data_tabel)


# ----------------------------------------------------
# FUNGSI – SPREADING OF CLAIM
# ----------------------------------------------------
def alokasikan_klaim(data_severitas, ur, layer, data_iterasi=None):
    hasil = []
    for idx, klaim in enumerate(data_severitas):
        row = {
            "Iterasi": data_iterasi.iloc[idx]["Iterasi"] if data_iterasi is not None else f"Real.{idx + 1}",
            "Severitas": int(klaim),
            "UR": 0,
        }
        for i in range(1, 7):
            row[f"Layer {i}"] = 0

        sisa = max(0, klaim)
        row["UR"] = min(sisa, ur)
        sisa -= row["UR"]

        for i, batas in enumerate(layer, start=1):
            if sisa <= 0:
                break
            row[f"Layer {i}"] = min(sisa, batas)
            sisa -= row[f"Layer {i}"]

        hasil.append(row)

    return pd.DataFrame(hasil)


# ----------------------------------------------------
# FUNGSI – RINGKASAN FREKUENSI & UR
# ----------------------------------------------------
def rangkum_berdasarkan_frekuensi(df_simulasi, df_soc, jumlah_iterasi):
    data_f = df_simulasi.dropna(subset=["Flagging Frekuensi"])[["Iterasi", "Flagging Frekuensi"]]
    peta_f = data_f.set_index("Iterasi")["Flagging Frekuensi"].to_dict()

    base_iter = df_simulasi["Iterasi"].str.split(".").str[0]
    ringkasan_sev = df_simulasi.groupby(base_iter)["Severitas"].sum().rename("Total Severitas")

    base_iter_soc = df_soc["Iterasi"].str.split(".").str[0]
    ringkasan_ur = df_soc.groupby(base_iter_soc)["UR"].sum().rename("Total UR")

    semua = [str(i) for i in range(1, jumlah_iterasi + 1)]

    df_out = pd.DataFrame(index=semua)
    df_out.index.name = "Iterasi"
    df_out["Frekuensi"] = [peta_f.get(f"{i}.1", 0) for i in semua]
    df_out["Total Severitas"] = [ringkasan_sev.get(i, 0) for i in semua]
    df_out["Total UR"] = [ringkasan_ur.get(i, 0) for i in semua]

    return df_out.reset_index()


# ----------------------------------------------------
# FUNGSI – RINGKASAN PER LAYER + REINSTATEMENT (SIMPLE)
# ----------------------------------------------------
def rangkum_layer(df_soc, nomor_layer, batas_layer, jumlah_iterasi, maks_reinst):
    base_iter = df_soc["Iterasi"].str.split(".").str[0]
    tot_layer = df_soc.groupby(base_iter)[f"Layer {nomor_layer}"].sum()

    semua = [str(i) for i in range(1, jumlah_iterasi + 1)]
    df_out = pd.DataFrame(index=semua)
    df_out.index.name = "Iterasi"

    df_out[f"Total Layer {nomor_layer}"] = [tot_layer.get(i, 0) for i in semua]

    # frekuensi: berapa klaim yang kena layer
    mask_pos = df_soc[f"Layer {nomor_layer}"] > 0
    freq_layer = df_soc[mask_pos].groupby(base_iter).size()
    df_out[f"Frekuensi Layer {nomor_layer}"] = [int(freq_layer.get(i, 0)) for i in semua]

    # premi reinstatement: incremental limit
    for r in range(maks_reinst + 1):
        cap_prev = r * batas_layer
        cap_now = (r + 1) * batas_layer
        pay_prev = np.minimum(df_out[f"Total Layer {nomor_layer}"], cap_prev)
        pay_now = np.minimum(df_out[f"Total Layer {nomor_layer}"], cap_now)
        incremental = pay_now - pay_prev
        df_out[f"Reinstatement {r}"] = incremental

    return df_out.reset_index()


# ----------------------------------------------------
# FUNGSI – HITUNG PREMI
# ----------------------------------------------------
def hitung_premi(daftar_df_layer, layer, reinstatement_per_layer):
    rows = []
    max_reinst_all = max(reinstatement_per_layer) if reinstatement_per_layer else 0

    for i, (df_layer, limit, max_reinst) in enumerate(
        zip(daftar_df_layer, layer, reinstatement_per_layer), start=1
    ):
        if limit <= 0 or df_layer.empty:
            continue

        col_total = f"Total Layer {i}"
        col_freq = f"Frekuensi Layer {i}"

        rata2_klaim = int(df_layer[col_total].mean())
        std_val = df_layer[col_total].std()
        std_klaim = int(std_val) if std_val == std_val else 0  # handle NaN
        frek = int(df_layer[col_freq].sum())
        total = int(df_layer[col_total].sum())

        row = {
            "Item": f"Layer {i}",
            "Batas": int(limit),
            "Rata-rata Klaim": rata2_klaim,
            "Standar Deviasi Klaim": std_klaim,
            "Frekuensi Klaim": frek,
            "Total Klaim": total,
        }

        for r in range(max_reinst_all + 1):
            if r > max_reinst:
                row[f"Reinstatement {r}"] = 0
            else:
                kolom = f"Reinstatement {r}"
                prem_r = int(df_layer[kolom].mean())
                row[kolom] = prem_r

        rows.append(row)

    if not rows:
        raise ValueError("Tidak ada layer yang aktif (batas = 0 semua).")

    df_premi = pd.DataFrame(rows)

    # baris total
    total_row = {
        "Item": "Total",
        "Batas": "",
        "Rata-rata Klaim": int(df_premi["Rata-rata Klaim"].sum()),
        "Standar Deviasi Klaim": int(df_premi["Standar Deviasi Klaim"].sum()),
        "Frekuensi Klaim": int(df_premi["Frekuensi Klaim"].sum()),
        "Total Klaim": int(df_premi["Total Klaim"].sum()),
    }
    for kol in df_premi.filter(like="Reinstatement").columns:
        total_row[kol] = int(df_premi[kol].sum())

    df_premi = pd.concat([df_premi, pd.DataFrame([total_row])], ignore_index=True)
    df_premi["Total"] = df_premi.filter(like="Reinstatement").sum(axis=1)

    return df_premi


# ----------------------------------------------------
# FUNGSI – RINGKASAN DATA ASLI
# ----------------------------------------------------
def ringkasan_data_asli(df_soc_real, ur, layer):
    summary = []

    ur_data = df_soc_real["UR"].fillna(0)
    tot_ur = int(ur_data.sum())
    freq_ur = int((ur_data > 0).sum())
    rata2_ur = int(tot_ur / freq_ur) if freq_ur > 0 else 0

    summary.append(
        {
            "Item": "OR",
            "Batas": int(ur),
            "Rata-rata Klaim (All Polis)": int(ur_data.mean()),
            "Frekuensi Klaim": freq_ur,
            "Total Klaim": tot_ur,
            "Rata-rata Klaim per OR/Layer": rata2_ur,
        }
    )

    for i in range(1, 7):
        data = df_soc_real[f"Layer {i}"].fillna(0)
        tot = int(data.sum())
        freq = int((data > 0).sum())
        rata2 = int(tot / freq) if freq > 0 else 0

        summary.append(
            {
                "Item": f"Layer {i}",
                "Batas": int(layer[i - 1]),
                "Rata-rata Klaim (All Polis)": int(data.mean()),
                "Frekuensi Klaim": freq,
                "Total Klaim": tot,
                "Rata-rata Klaim per OR/Layer": rata2,
            }
        )

    df_sum = pd.DataFrame(summary)
    total_row = {
        "Item": "Total",
        "Batas": "",
        "Rata-rata Klaim (All Polis)": int(df_sum["Rata-rata Klaim (All Polis)"].sum()),
        "Frekuensi Klaim": int(df_sum["Frekuensi Klaim"].sum()),
        "Total Klaim": int(df_sum["Total Klaim"].sum()),
        "Rata-rata Klaim per OR/Layer": int(
            df_sum["Total Klaim"].sum() / df_sum["Frekuensi Klaim"].sum()
        )
        if df_sum["Frekuensi Klaim"].sum() > 0
        else 0,
    }
    df_sum = pd.concat([df_sum, pd.DataFrame([total_row])], ignore_index=True)
    return df_sum


# ----------------------------------------------------
# UI – UNGGAH FILE
# ----------------------------------------------------
st.header("1. Unggah Data", divider="orange")
col1, col2 = st.columns(2)
with col1:
    file_sev = st.file_uploader("Unggah Data Severitas", type=["xlsx", "xls"], key="sev")
with col2:
    file_frek = st.file_uploader("Unggah Data Frekuensi", type=["xlsx", "xls"], key="frek")

if not (file_sev and file_frek):
    st.info("Silakan unggah kedua file (severitas & frekuensi) untuk memulai.")
    st.stop()

try:
    df_sev = pd.read_excel(file_sev)
    df_frek = pd.read_excel(file_frek)
except Exception as e:
    st.error(f"Gagal membaca file Excel: {e}")
    st.stop()

st.header("2. Pilih Kolom", divider="orange")
col1, col2 = st.columns(2)
with col1:
    kol_sev = st.selectbox("Kolom Severitas", df_sev.columns)
with col2:
    kol_frek = st.selectbox("Kolom Frekuensi", df_frek.columns)

data_sev = df_sev[kol_sev].dropna().values.astype(float)
data_frek = df_frek[kol_frek].dropna().values.astype(float)

if len(data_sev) == 0 or len(data_frek) == 0:
    st.error("Data severitas atau frekuensi kosong.")
    st.stop()

if np.any(data_sev <= 0):
    st.error("Data severitas harus > 0.")
    st.stop()

if np.any(data_frek < 0) or not np.all(data_frek == data_frek.astype(int)):
    st.error("Data frekuensi harus bilangan bulat ≥ 0.")
    st.stop()

# ----------------------------------------------------
# INPUT LAYER – DATA ASLI
# ----------------------------------------------------
st.header("3. Spreading of Claim – Data Asli", divider="orange")
st.caption("Gunakan OR & Layer berikut untuk membagi klaim historis.")

ur = rupiah_input("OR", key="ur_real", value=5_000_000_000, step=1_000_000_000)
layer_real = []
for i in range(1, 7):
    default = 0
    if i == 1:
        default = 5_000_000_000
    elif i == 2:
        default = 40_000_000_000
    elif i == 3:
        default = 50_000_000_000

    val = rupiah_input(f"Layer {i}", key=f"layer_real_{i}", value=default, step=1_000_000_000)
    layer_real.append(val)

# SoC data asli
df_soc_real = alokasikan_klaim(data_sev, ur, layer_real)

with st.expander(f"Lihat SoC (maks. {MAX_ROWS_DISPLAY} baris)", expanded=False):
    st.dataframe(df_soc_real.head(MAX_ROWS_DISPLAY), hide_index=True, use_container_width=True)
    if len(df_soc_real) > MAX_ROWS_DISPLAY:
        st.caption(f"Ditampilkan {MAX_ROWS_DISPLAY} baris pertama dari {len(df_soc_real):,}.")

df_summary = ringkasan_data_asli(df_soc_real, ur, layer_real)
st.subheader("Ringkasan Data Asli", divider="orange")
st.dataframe(df_summary, hide_index=True, use_container_width=True)

# ----------------------------------------------------
# PARAMETER FREKUENSI
# ----------------------------------------------------
st.header("4. Parameter Frekuensi", divider="orange")

mean_f = data_frek.mean()
var_f = data_frek.var()

param_poisson = {"mu": mean_f}
if var_f > mean_f:
    p_nb = mean_f / var_f
    n_nb = mean_f**2 / (var_f - mean_f)
else:
    p_nb = 0.9
    n_nb = max(mean_f, 1.0)
param_nb = {"p": float(p_nb), "n": float(n_nb)}
param_geom = {"p": 1 / mean_f if mean_f > 0 else 0.9}

c1, c2, c3 = st.columns(3)
with c1:
    st.write("**Poisson**")
    st.write(f"μ = {mean_f:.2f}")
with c2:
    st.write("**Negative Binomial**")
    st.write(f"p = {p_nb:.4f}")
    st.write(f"n = {n_nb:.2f}")
with c3:
    st.write("**Geometric**")
    st.write(f"p = {param_geom['p']:.4f}")

# ----------------------------------------------------
# INPUT SIMULASI
# ----------------------------------------------------
st.header("5. Pengaturan Simulasi Monte Carlo", divider="orange")
col1, col2 = st.columns(2)
with col1:
    jumlah_iterasi = st.number_input("Jumlah Iterasi", min_value=1, value=1_000, step=100)
with col2:
    dist_frek_pilih = st.selectbox("Distribusi Frekuensi", ["poisson", "nbinom", "geom"])

if jumlah_iterasi > 10_000:
    st.warning(
        "Jumlah iterasi besar bisa membuat proses lebih lama. "
        "Untuk dataset ~3.000 baris, 1.000–5.000 iterasi biasanya cukup."
    )

st.header("6. OR & Layer untuk Simulasi", divider="orange")
ur_sim = rupiah_input("OR (Simulasi)", key="ur_sim", value=5_000_000_000, step=1_000_000_000)

layer_sim = []
reins_sim = []
for i in range(1, 7):
    col_a, col_b = st.columns(2)
    with col_a:
        default = 0
        if i == 1:
            default = 5_000_000_000
        elif i == 2:
            default = 40_000_000_000
        elif i == 3:
            default = 50_000_000_000
        val = rupiah_input(
            f"Layer {i} (Simulasi)", key=f"layer_sim_{i}", value=default, step=1_000_000_000
        )
    with col_b:
        reins = st.number_input(
            f"Jumlah Reinstatement Layer {i}", min_value=0, max_value=10, value=4, step=1
        )
    layer_sim.append(val)
    reins_sim.append(reins)

# ----------------------------------------------------
# JALANKAN SIMULASI
# ----------------------------------------------------
st.header("7. Hasil Simulasi & Premi", divider="orange")

if st.button("Jalankan Simulasi", type="primary"):
    try:
        if dist_frek_pilih == "poisson":
            param_f = param_poisson
        elif dist_frek_pilih == "nbinom":
            param_f = param_nb
        else:
            param_f = param_geom

        # Simulasi klaim
        df_sim = simulasi_monte_carlo(jumlah_iterasi, dist_frek_pilih, param_f, data_sev)

        st.subheader("7.1 Hasil Simulasi Klaim", divider="orange")
        with st.expander(f"Lihat hasil simulasi (maks. {MAX_ROWS_DISPLAY} baris)", expanded=False):
            st.dataframe(df_sim.head(MAX_ROWS_DISPLAY), hide_index=True, use_container_width=True)
            if len(df_sim) > MAX_ROWS_DISPLAY:
                st.caption(f"Ditampilkan {MAX_ROWS_DISPLAY} baris pertama dari {len(df_sim):,}.")

        c1, c2 = st.columns(2)
        with c1:
            st.metric("Jumlah Baris Klaim", f"{len(df_sim):,}")
        with c2:
            st.metric("Rata-rata Severitas Simulasi", f"{df_sim['Severitas'].mean():,.0f}")

        # SoC simulasi
        df_soc_sim = alokasikan_klaim(df_sim["Severitas"].values, ur_sim, layer_sim, df_sim)

        st.subheader("7.2 Spreading of Claim (Simulasi)", divider="orange")
        with st.expander(f"Lihat SoC simulasi (maks. {MAX_ROWS_DISPLAY} baris)", expanded=False):
            st.dataframe(df_soc_sim.head(MAX_ROWS_DISPLAY), hide_index=True, use_container_width=True)
            if len(df_soc_sim) > MAX_ROWS_DISPLAY:
                st.caption(f"Ditampilkan {MAX_ROWS_DISPLAY} baris pertama dari {len(df_soc_sim):,}.")

        # Ringkasan UR
        df_ring_ur = rangkum_berdasarkan_frekuensi(df_sim, df_soc_sim, jumlah_iterasi)

        # Ringkasan per layer
        daftar_df_layer = []
        for i in range(1, 7):
            if layer_sim[i - 1] <= 0:
                df_empty = pd.DataFrame(
                    columns=["Iterasi", f"Total Layer {i}", f"Frekuensi Layer {i}", "Reinstatement 0"]
                )
                daftar_df_layer.append(df_empty)
                continue

            df_layer = rangkum_layer(
                df_soc_sim,
                nomor_layer=i,
                batas_layer=layer_sim[i - 1],
                jumlah_iterasi=jumlah_iterasi,
                maks_reinst=reins_sim[i - 1],
            )
            daftar_df_layer.append(df_layer)

            st.subheader(f"7.{2 + i} Ringkasan Layer {i}", divider="orange")
            with st.expander(f"Lihat ringkasan Layer {i}", expanded=False):
                st.dataframe(df_layer.drop(columns=["Iterasi"]), hide_index=True, use_container_width=True)

        # Premi XoL
        df_premi = hitung_premi(daftar_df_layer, layer_sim, reins_sim)
        st.subheader("10. Premi XoL", divider="orange")
        st.dataframe(df_premi, hide_index=True, use_container_width=True)

        # ------------------------------------------------
        # EXPORT KE EXCEL – HANYA 2 SHEET (RINGKASAN & PREMI)
        # ------------------------------------------------
        try:
            output = BytesIO()
            wb = Workbook()
            wb.remove(wb.active)

            border_thin = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            center = Alignment(horizontal="center", vertical="center")
            num_fmt = "#,##0"

            sheets = [
                (df_summary, "0. Ringkasan Data Klaim"),
                (df_premi, "1. Premi XoL"),
            ]

            for df_s, name in sheets:
                ws = wb.create_sheet(title=name)
                for r in dataframe_to_rows(df_s, index=False, header=True):
                    ws.append(r)

                # styling ringan
                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = border_thin
                        cell.alignment = center
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = num_fmt

                # auto width
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[col_letter].width = max_len + 4

            wb.save(output)
            output.seek(0)

            st.download_button(
                "Unduh Excel (Ringkasan & Premi)",
                data=output,
                file_name=f"Premi_XoL_Ringkasan_{dist_frek_pilih}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
        except Exception as e:
            st.error(f"Gagal membuat file Excel: {e}")

    except Exception as e:
        st.error(f"Terjadi error saat menjalankan simulasi: {e}")
